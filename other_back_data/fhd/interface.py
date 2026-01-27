from __future__ import annotations

import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from pydantic import BaseModel, Field, ConfigDict


def _safe_slug(s: str) -> str:
    s = (s or "").strip() or "default"
    s = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff._-]+", "-", s)
    return s[:120]


class FHDMaterializeInput(BaseModel):
    """Stable input schema for the FHD back-data adapter."""

    model_config = ConfigDict(extra="allow")

    output_dir: str = Field(..., description="Scenario output directory (e.g. outputs/<scenario_id>)")
    filters: Dict[str, Any] = Field(
        default_factory=dict,
        description=(
            "Filtering hints. Supported keys: province, city, district, park_name_contains, "
            "industry_keywords (list[str]), limit_parks (int)."
        ),
    )
    max_matched_rows: int = Field(default=5000, ge=0, le=200000)

    include_aoi_summary: bool = Field(default=True)
    aoi_compute_area_km2: bool = Field(
        default=False,
        description="If true, compute geodesic area (km2) for matched AOI polygons (may be slow).",
    )
    aoi_max_features_for_area: int = Field(default=2000, ge=0, le=200000)


class FHDMaterializeResult(BaseModel):
    """Structured output schema (JSON-friendly)."""

    model_config = ConfigDict(extra="allow")

    ok: bool
    source_id: str = "fhd"
    version: str = "0.1.0"
    metrics: Dict[str, Any] = Field(default_factory=dict)
    artifacts: Dict[str, Any] = Field(default_factory=dict)
    inventory_files: List[str] = Field(default_factory=list)
    recommended_inputs: Dict[str, Any] = Field(default_factory=dict)
    error: Optional[Dict[str, Any]] = None


def _discover_files(data_dir: Path) -> Tuple[Optional[Path], Optional[Path]]:
    excel = next(iter(sorted(data_dir.glob("*.xlsx"))), None)
    shp = next(iter(sorted(data_dir.glob("*.shp"))), None)
    return excel, shp


def _row_matches(row: Dict[str, Any], filters: Dict[str, Any]) -> bool:
    def contains(field: str, needle: str) -> bool:
        val = str(row.get(field) or "")
        return needle in val

    def fuzzy_contains(field: str, needle: str) -> bool:
        """More tolerant contains for place names.

        Some datasets may use '柳州' vs '柳州市'. We accept either direction.
        """

        val = str(row.get(field) or "").strip()
        n = str(needle or "").strip()
        if not n:
            return True
        if not val:
            return False
        return (n in val) or (val in n)

    province = (filters.get("province") or "").strip()
    if province and not fuzzy_contains("省份", province):
        return False

    city = (filters.get("city") or "").strip()
    if city and not fuzzy_contains("城市", city):
        return False

    district = (filters.get("district") or "").strip()
    if district and not fuzzy_contains("区县", district):
        return False

    park_name_contains = (filters.get("park_name_contains") or "").strip()
    if park_name_contains and not contains("产业园名称", park_name_contains):
        return False

    kws = filters.get("industry_keywords") or []
    if isinstance(kws, str):
        kws = [kws]
    kws = [str(k).strip() for k in kws if str(k).strip()]
    if kws:
        industry = str(row.get("产业") or "")
        if not any(k in industry for k in kws):
            return False

    return True


def _stream_excel_profile(
    excel_path: Path,
    filters: Dict[str, Any],
    max_matched_rows: int,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """Stream read the (large) park directory excel with openpyxl.

    Returns:
        metrics: aggregated counters + totals
        matched_rows: filtered rows (capped)
    """

    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header = [str(h).strip() if h is not None else "" for h in header]

    # expected headers
    # ['级别','产业','产业园名称','链接','简介','位置','经度','纬度','省份','城市','区县']

    total = 0
    matched: List[Dict[str, Any]] = []

    c_level: Counter[str] = Counter()
    c_industry: Counter[str] = Counter()
    c_province: Counter[str] = Counter()
    c_city: Counter[str] = Counter()

    c_matched_level: Counter[str] = Counter()
    c_matched_industry: Counter[str] = Counter()
    c_matched_province: Counter[str] = Counter()
    c_matched_city: Counter[str] = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        rec = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}

        lvl = str(rec.get("级别") or "").strip() or "未知"
        ind = str(rec.get("产业") or "").strip() or "未知"
        prov = str(rec.get("省份") or "").strip() or "未知"
        city = str(rec.get("城市") or "").strip() or "未知"

        c_level[lvl] += 1
        c_industry[ind] += 1
        c_province[prov] += 1
        c_city[city] += 1

        if _row_matches(rec, filters):
            if len(matched) < max_matched_rows:
                matched.append(rec)
            c_matched_level[lvl] += 1
            c_matched_industry[ind] += 1
            c_matched_province[prov] += 1
            c_matched_city[city] += 1

    wb.close()

    # Top lists (avoid huge dicts)
    metrics = {
        "total_parks": total,
        "level_distribution_top": c_level.most_common(20),
        "industry_distribution_top": c_industry.most_common(30),
        "province_distribution_top": c_province.most_common(30),
        "city_distribution_top": c_city.most_common(30),
        "matched_parks": len(matched),
        "matched_level_distribution_top": c_matched_level.most_common(20),
        "matched_industry_distribution_top": c_matched_industry.most_common(30),
        "matched_province_distribution_top": c_matched_province.most_common(30),
        "matched_city_distribution_top": c_matched_city.most_common(30),
    }
    return metrics, matched


def _write_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    # stable column order: use first row keys
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            # stringify to avoid Excel weirdness
            writer.writerow({k: ("" if v is None else str(v)) for k, v in r.items()})


def _aoi_summary(
    shp_path: Path,
    matched_park_names: Optional[set[str]] = None,
    compute_area_km2: bool = False,
    max_features_for_area: int = 2000,
) -> Dict[str, Any]:
    """Summarize AOI shapefile using fiona (avoid geopandas/pandas compatibility issues)."""

    import fiona  # type: ignore
    from shapely.geometry import shape  # type: ignore

    result: Dict[str, Any] = {
        "shp_path": str(shp_path),
        "crs": None,
        "schema": None,
        "total_features": None,
        "bounds": None,
        "type_distribution_top": [],
        "matched": {
            "matched_features": 0,
            "bounds": None,
            "area_km2": None,
        },
    }

    try:
        with fiona.open(shp_path) as src:
            result["crs"] = src.crs_wkt or (src.crs and str(src.crs))
            result["schema"] = src.schema
            try:
                result["total_features"] = len(src)
            except Exception:
                # fallback count
                result["total_features"] = sum(1 for _ in src)

            try:
                result["bounds"] = src.bounds
            except Exception:
                result["bounds"] = None

            c_type: Counter[str] = Counter()
            matched_bounds = None
            matched_area_m2 = 0.0
            matched_count = 0

            # pyproj optional area
            geod = None
            if compute_area_km2:
                try:
                    from pyproj import Geod  # type: ignore

                    geod = Geod(ellps="WGS84")
                except Exception:
                    geod = None

            for feat in src:
                props = feat.get("properties") or {}
                t = str(props.get("TYPE") or "")
                if t:
                    c_type[t] += 1

                if matched_park_names is not None:
                    name = str(props.get("NAME") or "")
                    if name and name in matched_park_names:
                        matched_count += 1
                        geom = None
                        try:
                            geom = shape(feat.get("geometry"))
                        except Exception:
                            geom = None
                        if geom is not None:
                            b = geom.bounds
                            if matched_bounds is None:
                                matched_bounds = list(b)
                            else:
                                matched_bounds[0] = min(matched_bounds[0], b[0])
                                matched_bounds[1] = min(matched_bounds[1], b[1])
                                matched_bounds[2] = max(matched_bounds[2], b[2])
                                matched_bounds[3] = max(matched_bounds[3], b[3])

                            if geod is not None and matched_count <= max_features_for_area:
                                try:
                                    area, _ = geod.geometry_area_perimeter(geom)
                                    matched_area_m2 += abs(float(area))
                                except Exception:
                                    pass

            result["type_distribution_top"] = c_type.most_common(20)
            result["matched"]["matched_features"] = matched_count
            result["matched"]["bounds"] = matched_bounds
            if geod is not None and matched_count and matched_count <= max_features_for_area:
                result["matched"]["area_km2"] = round(matched_area_m2 / 1e6, 4)

    except Exception as e:
        result["error"] = str(e)

    return result


def materialize(payload: FHDMaterializeInput) -> Dict[str, Any]:
    """Entry point used by multi_energy_agent.

    This function MUST NOT raise; it returns a structured dict.
    """

    out_dir = Path(payload.output_dir)
    artifacts_dir = out_dir / "artifacts"
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    data_dir = Path(__file__).resolve().parent
    excel_path, shp_path = _discover_files(data_dir)

    if not excel_path:
        return FHDMaterializeResult(
            ok=False,
            inventory_files=[],
            error={
                "type": "missing_file",
                "message": "FHD excel not found under other_back_data/fhd",
            },
        ).model_dump()

    inv = [str(excel_path)]
    if shp_path:
        inv.append(str(shp_path))

    try:
        metrics, matched_rows = _stream_excel_profile(
            excel_path=excel_path,
            filters=payload.filters,
            max_matched_rows=payload.max_matched_rows,
        )

        # Persist matched parks
        matched_csv = artifacts_dir / "fhd_matched_parks.csv"
        _write_csv(matched_rows, matched_csv)

        # Persist summary json
        summary_json = artifacts_dir / "fhd_summary.json"
        summary_json.write_text(json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8")

        # AOI summary
        aoi = None
        aoi_json = None
        if payload.include_aoi_summary and shp_path:
            matched_names = {str(r.get("产业园名称") or "").strip() for r in matched_rows if str(r.get("产业园名称") or "").strip()}
            aoi = _aoi_summary(
                shp_path,
                matched_park_names=matched_names if matched_names else None,
                compute_area_km2=payload.aoi_compute_area_km2,
                max_features_for_area=payload.aoi_max_features_for_area,
            )
            aoi_json = artifacts_dir / "fhd_aoi_summary.json"
            aoi_json.write_text(json.dumps(aoi, ensure_ascii=False, indent=2), encoding="utf-8")

        # Recommended inputs: we intentionally keep them small.
        # The original excel (100k+ rows) is too heavy to re-ingest repeatedly.
        recommended_inputs = {
            "csv_paths": [str(matched_csv)],
            "json_paths": [str(summary_json)] + ([str(aoi_json)] if aoi_json else []),
        }

        return FHDMaterializeResult(
            ok=True,
            metrics=metrics,
            artifacts={
                "filters": payload.filters,
                "excel_path": str(excel_path),
                "shp_path": str(shp_path) if shp_path else None,
                "summary_json": str(summary_json),
                "matched_parks_csv": str(matched_csv),
                "aoi_summary_json": str(aoi_json) if aoi_json else None,
                "aoi_summary": aoi,
            },
            inventory_files=inv,
            recommended_inputs=recommended_inputs,
        ).model_dump()

    except Exception as e:  # never raise
        return FHDMaterializeResult(
            ok=False,
            inventory_files=inv,
            error={"type": "exception", "message": str(e)},
        ).model_dump()
